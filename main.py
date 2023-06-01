import openpyxl


#Abre os arquivos xlsx
workbookMERCOS = openpyxl.load_workbook("C:\\Users\\Usefr\\Desktop\\SincronizaPreçoB2B\\MERCOS.xlsx")
workbookFASTCOMMERCE = openpyxl.load_workbook("C:\\Users\\Usefr\\Desktop\\SincronizaPreçoB2B\\FASTCOMMERCE.xlsx")

#Cria o arquivo de saida
outputworkbook = openpyxl.Workbook()
outputworkbook2 = openpyxl.Workbook()

#Abre as planilhas
sheetMERCOS = workbookMERCOS.active
sheetFASTCOMMERCE = workbookFASTCOMMERCE.active
outputsheet = outputworkbook.active
outputsheet2 = outputworkbook2.active

#Contagem das filas
rownumber = 1
rownumber2 = 1

#Dicionarios que armazenam as informações dos arquivos
dictMERCOS = {}
dictFASTCOMMERCE = {}

for row in sheetMERCOS.iter_rows(values_only=True):
    dictMERCOS[row[0]] = row[1]

for row in sheetFASTCOMMERCE.iter_rows(values_only=True):
    dictFASTCOMMERCE[row[0]] = [row[1], row[2]]

#Cria o arquivo de saida com as informações necessárias
for key in dictFASTCOMMERCE:
    if key in dictMERCOS:
        values = dictFASTCOMMERCE[key]
        value1 = values[0]
        outputsheet.cell(row=rownumber, column=1, value=value1)
        outputsheet.cell(row=rownumber, column=2, value=dictMERCOS[key])
        rownumber += 1
    else:
        values = dictFASTCOMMERCE[key]
        value2 = values[1]
        outputsheet2.cell(row=rownumber2, column=1, value=key)
        outputsheet2.cell(row=rownumber2, column=2, value=value2)
        rownumber2 += 1

outputworkbook.save('C:\\Users\\Usefr\\Desktop\\SincronizaPreçoB2B\\Output.xlsx')
outputworkbook2.save('C:\\Users\\Usefr\\Desktop\\SincronizaPreçoB2B\\Output2.xlsx')